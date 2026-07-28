"""Monthly Report — audience: management, and (with audience=client) the
client themselves.

Answers "what is the overall shape of this project". The reader is furthest
from the work, so this summarises rather than groups: one executive page,
then a small number of roll-ups with charts. No line-by-line log at all —
that is the Daily's job.

The audience switch is a disclosure boundary, not a formatting preference.
`audience=client` must not emit per-person utilization, margin or rate,
delivery mode, or real names. Everything that could leak is gated in ONE
place — `_is_client()` — and each gated block says what it withholds and why.
"""

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from sqlalchemy.orm import Session

from .. import effort_budget, models, progress_matrix
from ..routers.projects import MANDATORY_COLUMN_BY_CATEGORY
from . import activity_feed
from .style import add_chart, write_empty_notice, write_note, write_section, write_table, write_title

AUDIENCES = ("internal", "client")

SUMMARY_COLUMNS = ["metric", "value"]
PHASE_COLUMNS = [
    "phase",
    "functions_total",
    "functions_done",
    "functions_pct",
    "tasks_total",
    "tasks_done",
    "tasks_pct",
    "documents_confirmed",
    "documents_total",
    "delay_days_avg",
]
BUDGET_COLUMNS = ["metric", "man_days"]
CR_COLUMNS = ["cr_code", "title", "status", "effort_md", "approved_on"]
CR_COLUMNS_INTERNAL = CR_COLUMNS + ["estimated_cost_thb"]
HEALTH_COLUMNS = ["state", "items", "percent"]
PHASE_DELAY_COLUMNS = ["phase", "items", "avg_delay_days", "worst_delay_days"]
RISK_COLUMNS = ["module", "entity_code", "entity_title", "owner", "state", "days_late", "suggested_action"]
UTILIZATION_COLUMNS = ["resource", "avg_percent", "peak_percent"]

FUNCTION_DONE_STATUSES = ("Confirmed", "Done")
TASK_DONE_STATUSES = ("Done",)
DOCUMENT_DONE_STATUSES = ("Confirmed",)

HEALTH_LABELS = {
    "on_track": "On track",
    "late": "Late",
    "overdue": "Overdue",
    "not_started_late": "Late to start",
    "unplanned": "No plan dates",
}


def _is_client(audience: str) -> bool:
    return audience == "client"


def _pct(done: int, total: int) -> str:
    return "-" if total == 0 else f"{round(done / total * 100)}%"


def generate(
    slug: str,
    year: int,
    month: int,
    db: Session,
    master_db: Session,
    audience: str = "internal",
) -> Workbook:
    client = _is_client(audience)
    month_start = date(year, month, 1)
    month_end = date(year + (month == 12), (month % 12) + 1, 1)
    start_dt, end_dt = activity_feed.range_bounds(month_start, month_end)

    project = master_db.query(models.Project).filter(models.Project.slug == slug).first()
    functions = db.query(models.Function).all()
    tasks = db.query(models.Task).all()
    documents = db.query(models.Document).all()
    activity = activity_feed.fetch(db, start_dt, end_dt)
    status_rows = activity_feed.status_changes(activity)
    matrix = progress_matrix.build_progress_matrix(
        slug=slug, db=db, master_db=master_db, entity_types=list(progress_matrix.ENTITY_MODELS)
    )
    budget = effort_budget.compute_effort_budget(db)
    config = effort_budget.get_config(db)
    roles = activity_feed.role_map(master_db)

    wb = Workbook()

    # ---- 1. Executive Summary ----
    ws = wb.active
    ws.title = "Executive Summary"
    row = write_title(ws, 1, f"Monthly Report — {project.name if project else slug} — {month_start.strftime('%B %Y')}")
    row = write_note(
        ws,
        row,
        "Client copy — commercial and staffing detail withheld."
        if client
        else "Internal copy — includes staffing and commercial detail.",
    )

    total_items = len(functions) + len(tasks)
    done_items = sum(1 for f in functions if f.status in FUNCTION_DONE_STATUSES) + sum(
        1 for t in tasks if t.status in TASK_DONE_STATUSES
    )
    health_counts: dict = {}
    for r in matrix["rows"]:
        health_counts[r["health"]] = health_counts.get(r["health"], 0) + 1
    approved_this_month = [
        r for r in status_rows if r["entity_type"] == "change_request" and r["to"] == "Approved"
    ]

    metrics = [
        {"metric": "Overall completion", "value": _pct(done_items, total_items)},
        {"metric": "Functions complete", "value": f"{sum(1 for f in functions if f.status in FUNCTION_DONE_STATUSES)} / {len(functions)}"},
        {"metric": "Tasks complete", "value": f"{sum(1 for t in tasks if t.status in TASK_DONE_STATUSES)} / {len(tasks)}"},
        {"metric": "Documents confirmed", "value": f"{sum(1 for d in documents if d.status in DOCUMENT_DONE_STATUSES)} / {len(documents)}"},
        {"metric": "On track", "value": health_counts.get("on_track", 0)},
        {"metric": "Late / overdue", "value": health_counts.get("late", 0) + health_counts.get("overdue", 0) + health_counts.get("not_started_late", 0)},
        {"metric": "Change requests approved this month", "value": len(approved_this_month)},
        {"metric": "Changes recorded this month", "value": len(activity)},
    ]
    if budget["contracted_md"] is not None:
        metrics.append({"metric": "Effort remaining", "value": f"{budget['remaining_md']:.1f} / {budget['contracted_md']:.0f} MD"})
    row = write_section(ws, row, "At a glance")
    write_table(ws, row, SUMMARY_COLUMNS, metrics)

    # ---- 2. Phase Breakdown ----
    ws2 = wb.create_sheet("Phase Breakdown")
    row = write_title(ws2, 1, "Progress by phase")
    phases = sorted(
        {f.phase for f in functions if f.phase} | {t.phase for t in tasks if t.phase} | {d.phase for d in documents if d.phase},
        key=models.phase_sort_key,
    )
    delays_by_phase: dict = {}
    for r in matrix["rows"]:
        if r["end_delay_days"]:
            delays_by_phase.setdefault(r["phase"], []).append(r["end_delay_days"])

    phase_rows = []
    for phase in phases:
        f_in = [f for f in functions if f.phase == phase]
        t_in = [t for t in tasks if t.phase == phase]
        d_in = [d for d in documents if d.phase == phase]
        f_done = sum(1 for f in f_in if f.status in FUNCTION_DONE_STATUSES)
        t_done = sum(1 for t in t_in if t.status in TASK_DONE_STATUSES)
        delays = delays_by_phase.get(phase, [])
        phase_rows.append(
            {
                "phase": phase,
                "functions_total": len(f_in),
                "functions_done": f_done,
                "functions_pct": _pct(f_done, len(f_in)),
                "tasks_total": len(t_in),
                "tasks_done": t_done,
                "tasks_pct": _pct(t_done, len(t_in)),
                "documents_confirmed": sum(1 for d in d_in if d.status in DOCUMENT_DONE_STATUSES),
                "documents_total": len(d_in),
                "delay_days_avg": round(sum(delays) / len(delays), 1) if delays else "",
            }
        )
    if phase_rows:
        header_row = row
        row = write_table(ws2, row, PHASE_COLUMNS, phase_rows)
        pct_col = len(PHASE_COLUMNS) + 2
        ws2.cell(row=header_row, column=pct_col, value="overall_pct")
        for i, r in enumerate(phase_rows):
            total = r["functions_total"] + r["tasks_total"]
            done = r["functions_done"] + r["tasks_done"]
            ws2.cell(row=header_row + 1 + i, column=pct_col, value=round(done / total * 100) if total else 0)
        chart = BarChart()
        chart.title = "% completion by phase"
        chart.y_axis.title = "%"
        add_chart(
            ws2,
            chart,
            f"A{row + 1}",
            Reference(ws2, min_col=pct_col, min_row=header_row, max_row=header_row + len(phase_rows)),
            Reference(ws2, min_col=1, min_row=header_row + 1, max_row=header_row + len(phase_rows)),
            len(phase_rows),
            skip_note_cell=f"A{row + 1}",
            reason="Completion chart omitted — needs at least 2 phases carrying items.",
        )
    else:
        write_empty_notice(ws2, row, PHASE_COLUMNS, "No functions, tasks or documents carry a phase yet.")

    # ---- 3. Effort & Budget ----
    ws3 = wb.create_sheet("Effort & Budget")
    row = write_title(ws3, 1, "Effort against contract")
    if budget["contracted_md"] is None:
        write_empty_notice(
            ws3, row, BUDGET_COLUMNS,
            "No contracted man-days are recorded for this project, so there is no budget to report against.",
        )
    else:
        budget_rows = [
            {"metric": "Contracted", "man_days": round(budget["contracted_md"], 1)},
            {"metric": "Used (delivered)", "man_days": round(budget["used_md"], 1)},
            {"metric": "Committed (in flight)", "man_days": round(budget["committed_md"], 1)},
            {"metric": "Remaining", "man_days": round(budget["remaining_md"], 1)},
        ]
        header_row = row
        row = write_table(ws3, row, BUDGET_COLUMNS, budget_rows)
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Effort budget"
        chart.y_axis.title = "man-days"
        add_chart(
            ws3,
            chart,
            f"A{row + 1}",
            Reference(ws3, min_col=2, min_row=header_row, max_row=header_row + 3),
            Reference(ws3, min_col=1, min_row=header_row + 1, max_row=header_row + 3),
            len(budget_rows),
            skip_note_cell=f"A{row + 1}",
        )
        row += 16
        # Rate and margin are internal-only. The client gets the man-days,
        # which is what the contract is denominated in.
        if not client and config.rate_thb_per_md:
            row = write_section(ws3, row, "Commercial (internal only)")
            row = write_table(
                ws3,
                row,
                ["metric", "value"],
                [
                    {"metric": "Day rate (THB/MD)", "value": config.rate_thb_per_md},
                    {"metric": "Used, at rate", "value": round(budget["used_md"] * config.rate_thb_per_md)},
                    {"metric": "Remaining, at rate", "value": round(budget["remaining_md"] * config.rate_thb_per_md)},
                ],
            )
        # Delivery mode obeys the existing per-project flag as well as the
        # audience — a client copy shows it only when the project has
        # explicitly opted in.
        show_delivery = bool(config.show_delivery_mode_in_client_docs) if client else True
        if show_delivery:
            modes = {
                (e.delivery_mode or "human")
                for e in db.query(models.EffortEstimate).all()
            }
            if modes:
                row = write_section(ws3, row, "Delivery model")
                write_table(
                    ws3,
                    row,
                    ["metric", "value"],
                    [{"metric": "Modes in use", "value": ", ".join(sorted(modes))}],
                )

    # ---- 4. Change Request Summary ----
    ws4 = wb.create_sheet("Change Requests")
    row = write_title(ws4, 1, "Change requests")
    crs = db.query(models.ChangeRequest).order_by(models.ChangeRequest.id).all()
    columns = CR_COLUMNS if client else CR_COLUMNS_INTERNAL
    if crs:
        estimates: dict = {}
        for e in db.query(models.EffortEstimate).filter(
            models.EffortEstimate.linked_entity_type == "change_request"
        ).all():
            estimates[e.linked_entity_id] = estimates.get(e.linked_entity_id, 0) + (e.calculated_man_days or 0)
        approved_dates = {
            r["entity_id"]: r["datetime"] for r in status_rows
            if r["entity_type"] == "change_request" and r["to"] == "Approved"
        }
        rows = []
        for c in crs:
            entry = {
                "cr_code": c.cr_code,
                "title": c.title,
                "status": c.status,
                "effort_md": round(estimates.get(c.id, 0), 2) or "",
                "approved_on": approved_dates.get(c.id, ""),
            }
            if not client:
                md = estimates.get(c.id, 0)
                entry["estimated_cost_thb"] = round(md * config.rate_thb_per_md) if config.rate_thb_per_md else ""
            rows.append(entry)
        write_table(ws4, row, columns, rows)
    else:
        write_empty_notice(ws4, row, columns, "No change requests have been raised on this project.")

    # ---- 5. Schedule Health ----
    ws5 = wb.create_sheet("Schedule Health")
    row = write_title(ws5, 1, "Schedule health")
    total_rows = len(matrix["rows"])
    health_rows = [
        {
            "state": HEALTH_LABELS.get(state, state),
            "items": count,
            "percent": f"{round(count / total_rows * 100)}%" if total_rows else "-",
        }
        for state, count in sorted(health_counts.items(), key=lambda kv: -kv[1])
    ]
    if health_rows:
        header_row = row
        row = write_table(ws5, row, HEALTH_COLUMNS, health_rows)
        chart = PieChart()
        chart.title = "Items by schedule state"
        add_chart(
            ws5,
            chart,
            f"A{row + 1}",
            Reference(ws5, min_col=2, min_row=header_row, max_row=header_row + len(health_rows)),
            Reference(ws5, min_col=1, min_row=header_row + 1, max_row=header_row + len(health_rows)),
            len(health_rows),
            skip_note_cell=f"A{row + 1}",
            reason="Schedule pie omitted — needs at least 2 different states to compare.",
        )
        row += 16
    else:
        row = write_empty_notice(ws5, row, HEALTH_COLUMNS, "No items are being tracked in the Progress Matrix yet.")

    row = write_section(ws5, row, "Delay by phase")
    delay_rows = [
        {
            "phase": phase or "(no phase)",
            "items": len(delays),
            "avg_delay_days": round(sum(delays) / len(delays), 1),
            "worst_delay_days": max(delays),
        }
        for phase, delays in sorted(delays_by_phase.items(), key=lambda kv: models.phase_sort_key(kv[0]))
    ]
    if delay_rows:
        write_table(ws5, row, PHASE_DELAY_COLUMNS, delay_rows)
    else:
        write_empty_notice(ws5, row, PHASE_DELAY_COLUMNS, "Nothing has finished late, so there is no delay to break down.")

    # ---- 6. Risk & Recovery ----
    ws6 = wb.create_sheet("Risk & Recovery")
    row = write_title(ws6, 1, "At-risk items and suggested recovery")
    risk_rows = [
        {
            "module": r.get("item_type") or r["entity_type"],
            "entity_code": r["entity_code"] or f"#{r['entity_id']}",
            "entity_title": r["entity_title"],
            # Client copies get the role, never the person.
            "owner": activity_feed.to_role(r["owner"], roles) if client else r["owner"],
            "state": HEALTH_LABELS.get(r["health"], r["health"]),
            "days_late": r["end_delay_days"] if r["end_delay_days"] is not None else r["start_delay_days"],
            "suggested_action": _scrub(r["recovery"][0]["action"], roles) if (client and r["recovery"]) else (r["recovery"][0]["action"] if r["recovery"] else ""),
        }
        for r in matrix["rows"]
        if r["health"] in ("overdue", "not_started_late", "late")
    ]
    risk_rows.sort(key=lambda r: (r["days_late"] is None, -(r["days_late"] or 0)))
    if risk_rows:
        write_table(ws6, row, RISK_COLUMNS, risk_rows)
    else:
        write_empty_notice(ws6, row, RISK_COLUMNS, "Nothing is currently flagged at risk.")

    # ---- 7. Resource Utilization — internal only ----
    if client:
        # The sheet is omitted entirely rather than emptied, so there is no
        # header hinting at what was removed.
        pass
    else:
        ws7 = wb.create_sheet("Resource Utilization")
        row = write_title(ws7, 1, "Resource utilization this month")
        row = write_note(ws7, row, "Internal only — omitted entirely from client copies.")
        util = _utilization(master_db, month_start, month_end)
        if util:
            write_table(ws7, row, UTILIZATION_COLUMNS, util)
        else:
            write_empty_notice(ws7, row, UTILIZATION_COLUMNS, "No resources are allocated over this month.")

    return wb


def _scrub(text: str, roles: dict) -> str:
    """Replaces any resource name appearing in free text with their role.

    Recovery suggestions embed the owner's name in their wording ("Bee is
    allocated 80%..."), so scrubbing the owner column alone would still leak
    it here.
    """
    if not text:
        return text
    out = text
    # Longest first, so "Anna Marie" is replaced before "Anna".
    for name in sorted(roles, key=len, reverse=True):
        if name and name in out:
            out = out.replace(name, roles[name])
    return out


def _utilization(master_db: Session, start: date, end: date) -> list[dict]:
    from ..resource_utils import compute_utilization

    rows = []
    for r in compute_utilization(master_db, start, end - timedelta(days=1)):
        percents = [w.get("total_percent") or 0 for w in (r.get("weeks") or [])]
        if not percents or not any(percents):
            continue
        rows.append(
            {
                "resource": r["resource_name"],
                "avg_percent": round(sum(percents) / len(percents)),
                "peak_percent": max(percents),
            }
        )
    return rows

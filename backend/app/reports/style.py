import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri")
TITLE_FONT = Font(name="Calibri", bold=True, size=13)
SECTION_FONT = Font(name="Calibri", bold=True, size=11)


def write_title(ws, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    return row + 2


def write_section(ws, row: int, text: str) -> int:
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    return row + 1


def write_table(ws, start_row: int, headers: list[str], rows: list[dict]) -> int:
    """Writes a styled header row followed by data rows starting at
    `start_row`. Returns the next free row after the table."""
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    r = start_row + 1
    for row in rows:
        for c, h in enumerate(headers, start=1):
            value = row.get(h)
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT
        r += 1

    for c, h in enumerate(headers, start=1):
        col_letter = get_column_letter(c)
        width = max(len(h) + 2, 12)
        ws.column_dimensions[col_letter].width = min(width, 40)

    return r + 1


NOTE_FONT = Font(name="Calibri", italic=True, color="6B7280")

# A chart needs at least this many data points to say anything. Below it, the
# chart is skipped and a note explains why — an axis with one bar on it looks
# like a broken report, and an empty one looks like broken data.
MIN_CHART_POINTS = 2


def write_note(ws, row: int, text: str) -> int:
    """An explanatory line, for the cases where there is nothing to show."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    return row + 2


def write_empty_notice(ws, row: int, headers: list[str], reason: str) -> int:
    """A sheet with no data still gets its header row, so the reader can see
    what it *would* contain, plus a sentence saying why it is empty. A blank
    sheet is indistinguishable from a broken one."""
    row = write_table(ws, row, headers, [])
    return write_note(ws, row, reason)


def add_chart(ws, chart, anchor: str, data_ref, cats_ref, point_count: int, skip_note_cell: str = None, reason: str = None):
    """Attaches a native Excel chart bound to cell ranges in this sheet, or
    skips it and writes a note when there isn't enough data to plot.

    Native charts (not images) so the reader can retune the range, restyle it,
    or copy it into a deck — and so it updates if they edit the numbers.
    """
    if point_count < MIN_CHART_POINTS:
        if skip_note_cell:
            cell = ws[skip_note_cell]
            cell.value = reason or (
                f"Chart omitted — needs at least {MIN_CHART_POINTS} data points, found {point_count}."
            )
            cell.font = NOTE_FONT
        return None
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor)
    return chart


def workbook_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

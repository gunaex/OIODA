"""Template generation, import validation and export, driven by
`import_schemas`.

The three used to live in five different routers with their own hardcoded
column lists. They now share one definition, so a new model field cannot end
up in export but not the template, or in the template but rejected on import.

Validation reports every problem it finds with its row number, rather than
stopping at the first — importing a 200-row spreadsheet and being told about
one bad cell at a time is not a workable way to fix a file.
"""

import io
from datetime import date, datetime
from typing import Any

import pandas as pd
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .import_schemas import LEGACY_PHASE_MAP, ImportSchema

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
NOTE_FONT = Font(name="Calibri", italic=True, color="6B7280")

# Excel refuses an inline dropdown list longer than 255 characters. Longer
# enumerations fall back to being documented on the Instructions sheet and
# validated server-side, which is the real gate anyway.
MAX_INLINE_VALIDATION_CHARS = 255

# The first data row in a generated template (row 1 is the header).
FIRST_DATA_ROW = 2
LAST_VALIDATED_ROW = 1000


# --------------------------------------------------------------------------
# Template generation
# --------------------------------------------------------------------------


def build_template(schema: ImportSchema, columns: list = None) -> Workbook:
    """A blank template: one header row, a dropdown on every enum column, and
    an Instructions sheet explaining the columns that are computed rather than
    entered."""
    columns = columns or schema.template_columns
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for index, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=index, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(index)].width = min(max(len(name) + 4, 12), 34)

        allowed = schema.enums.get(name)
        if allowed:
            formula = '"' + ",".join(str(v) for v in allowed) + '"'
            if len(formula) <= MAX_INLINE_VALIDATION_CHARS:
                dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
                dv.error = f"{name} must be one of: {', '.join(str(v) for v in allowed)}"
                dv.errorTitle = "Value not allowed"
                dv.prompt = f"Allowed: {', '.join(str(v) for v in allowed)}"
                ws.add_data_validation(dv)
                letter = get_column_letter(index)
                dv.add(f"{letter}{FIRST_DATA_ROW}:{letter}{LAST_VALIDATED_ROW}")

    ws.freeze_panes = "A2"

    # ---- Instructions ----
    info = wb.create_sheet("Instructions")
    info.column_dimensions["A"].width = 34
    info.column_dimensions["B"].width = 90
    row = 1
    info.cell(row=row, column=1, value="Column").font = HEADER_FONT
    info.cell(row=row, column=1).fill = HEADER_FILL
    info.cell(row=row, column=2, value="Notes").font = HEADER_FONT
    info.cell(row=row, column=2).fill = HEADER_FILL
    row += 1

    for name in columns:
        parts = []
        if name in schema.required:
            parts.append("REQUIRED.")
        allowed = schema.enums.get(name)
        if allowed:
            parts.append("One of: " + ", ".join(str(v) for v in allowed) + ".")
        if name in schema.notes:
            parts.append(schema.notes[name])
        info.cell(row=row, column=1, value=name)
        info.cell(row=row, column=2, value=" ".join(parts) or "Free text.")
        row += 1

    if schema.export_only or schema.derived_export:
        row += 1
        info.cell(row=row, column=1, value="Computed / not imported").font = HEADER_FONT
        info.cell(row=row, column=1).fill = HEADER_FILL
        row += 1
        for name in list(schema.export_only) + list(schema.derived_export):
            info.cell(row=row, column=1, value=name)
            info.cell(
                row=row,
                column=2,
                value=schema.notes.get(name, "Appears on export for reference. Ignored if present on import."),
            ).font = NOTE_FONT
            row += 1

    if LEGACY_PHASE_MAP and "phase" in columns:
        row += 1
        info.cell(row=row, column=1, value="Legacy phase values").font = HEADER_FONT
        info.cell(row=row, column=1).fill = HEADER_FILL
        row += 1
        info.cell(
            row=row,
            column=2,
            value="Older files are accepted and converted: "
            + ", ".join(f"{old} -> {new}" for old, new in LEGACY_PHASE_MAP.items())
            + ". The conversions are listed in the import result.",
        ).font = NOTE_FONT

    return wb


def template_response(schema: ImportSchema, filename: str, columns: list = None) -> StreamingResponse:
    buf = io.BytesIO()
    build_template(schema, columns).save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------------------------------------------------------------
# Import parsing + validation
# --------------------------------------------------------------------------


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.date()
    if value is not None and not isinstance(value, (str, int, float, bool, date, datetime)):
        return str(value)
    return value


DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d")


def _coerce_date(value: Any):
    """A date cell can arrive as a real date, a Timestamp, or text — Excel
    doesn't distinguish for the person filling it in, so all three are
    accepted. Returns None when it genuinely isn't a date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _coerce_bool(value: Any) -> Any:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in ("true", "yes", "y", "1"):
            return True
        if lowered in ("false", "no", "n", "0"):
            return False
    return value


class ImportError_(Exception):
    def __init__(self, errors: list, unknown_columns: list = None):
        self.errors = errors
        self.unknown_columns = unknown_columns or []


def read_rows(file_bytes: bytes, schema: ImportSchema, columns: list = None) -> tuple:
    """Parses and validates a spreadsheet against a schema.

    Returns (records, report). Raises ImportError_ carrying EVERY problem
    found — with the row number as it appears in Excel — rather than failing
    on the first one.
    """
    columns = columns or schema.template_columns
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}")

    actual = [str(c).strip() for c in df.columns]
    missing = [c for c in columns if c not in actual]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Column headers do not match the import template.",
                "missing_columns": missing,
                "expected_columns": columns,
                "found_columns": actual,
            },
        )

    # Columns the schema knows but doesn't import are skipped in silence —
    # that is what makes an exported file re-importable. Anything else is
    # surfaced, because it is usually a typo in a header.
    unknown = [c for c in actual if not schema.known(c)]
    ignored_known = [c for c in actual if c not in columns and schema.known(c)]

    df = df[columns].astype(object).where(df[columns].notnull(), None)
    raw_records = df.to_dict(orient="records")

    errors: list = []
    conversions: list = []
    records: list = []

    for index, raw in enumerate(raw_records):
        # +2: one for the header row, one because spreadsheets are 1-based.
        excel_row = index + 2
        record = {key: _normalize_cell(value) for key, value in raw.items()}

        if all(v in (None, "") for v in record.values()):
            continue  # a blank spacer row is not an error

        for column in columns:
            value = record.get(column)
            if isinstance(value, str):
                value = value.strip()
                record[column] = value or None
                value = record[column]

            if column in schema.required and value in (None, ""):
                errors.append({"row": excel_row, "column": column, "value": None, "problem": "required value is missing"})
                continue

            if value in (None, ""):
                continue

            if column in schema.dates:
                parsed = _coerce_date(value)
                if parsed is None:
                    errors.append(
                        {
                            "row": excel_row,
                            "column": column,
                            "value": str(value),
                            "problem": "is not a date the sheet could be read as — use a real date cell, "
                            "or type it as YYYY-MM-DD",
                        }
                    )
                else:
                    record[column] = parsed
                continue

            allowed = schema.enums.get(column)
            if not allowed:
                continue

            if allowed == ("TRUE", "FALSE"):
                record[column] = _coerce_bool(value)
                continue

            text = str(value).strip()
            if text in allowed:
                record[column] = text
                continue

            # Legacy phase names are converted rather than rejected, so a file
            # the team already has imports without hand-editing. Every
            # conversion is reported back.
            if column == "phase" and text in LEGACY_PHASE_MAP:
                new_value = LEGACY_PHASE_MAP[text]
                record[column] = new_value
                conversions.append({"row": excel_row, "column": column, "from": text, "to": new_value})
                continue

            errors.append(
                {
                    "row": excel_row,
                    "column": column,
                    "value": text,
                    "problem": f"not a recognised value — allowed: {', '.join(str(v) for v in allowed)}"
                    + (
                        f" (legacy values also accepted: {', '.join(LEGACY_PHASE_MAP)})"
                        if column == "phase"
                        else ""
                    ),
                }
            )

        records.append(record)

    if errors:
        raise ImportError_(errors, unknown)

    return records, {
        "converted": conversions,
        "unknown_columns": unknown,
        "ignored_columns": ignored_known,
    }


def raise_import_errors(exc: ImportError_):
    """Turns a validation failure into a 400 the UI can render as a list."""
    raise HTTPException(
        status_code=400,
        detail={
            "message": f"{len(exc.errors)} cell(s) could not be imported. Nothing was saved.",
            "errors": exc.errors[:200],
            "error_count": len(exc.errors),
            "unknown_columns": exc.unknown_columns,
        },
    )


# --------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------


def export_response(rows: list, columns: list, filename: str) -> StreamingResponse:
    df = pd.DataFrame(rows, columns=columns)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

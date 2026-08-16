"""Daily Report — audience: the PM and the delivery team, read same-day.

Answers "what moved today, and what needs looking at tomorrow morning". That
makes it the detailed one of the three: every log entry is listed line by
line, because the reader is close enough to the work to want the detail. The
Weekly groups instead of listing, and the Monthly summarises instead of
grouping — see the audience table in the spec.
"""

from datetime import date

from openpyxl import Workbook
from sqlalchemy.orm import Session

from .. import models, progress_matrix
from . import activity_feed
from .style import write_empty_notice, write_note, write_section, write_table, write_title

SUMMARY_COLUMNS = ["module", "changes_today", "status_changes"]
ACTIVITY_COLUMNS = ["time", "by", "module", "entity_code", "entity_title", "field", "from", "to"]
STATUS_COLUMNS = ["time", "module", "entity_code", "entity_title", "from", "to", "by"]
NOTE_COLUMNS = ["source", "ref", "title_or_content", "created_by", "created_at"]
ATTENTION_COLUMNS = [
    "module",
    "entity_code",
    "entity_title",
    "owner",
    "state",
    "plan_end",
    "days_late",
    "suggested_action",
]


def _notes_created(db: Session, start_dt, end_dt) -> list[dict]:
    """Both note tables. The quick-capture `notes` table was the only one the
    original report knew about; `note_pages` (the markdown wiki added later)
    was invisible here until now."""
    rows = []
    for n in (
        db.query(models.Note)
        .filter(models.Note.created_at >= start_dt, models.Note.created_at < end_dt)
        .order_by(models.Note.created_at)
        .all()
    ):
        rows.append(
            {
                "source": "Quick Note",
                "ref": f"#{n.id}",
                "title_or_content": n.content,
                "created_by": None,
                "created_at": n.created_at.strftime("%Y-%m-%d %H:%M"),
            }
        )
    for p in (
        db.query(models.NotePage)
        .filter(models.NotePage.created_at >= start_dt, models.NotePage.created_at < end_dt)
        .order_by(models.NotePage.created_at)
        .all()
    ):
        rows.append(
            {
                "source": "Note Page",
                "ref": f"#{p.id}",
                "title_or_content": p.title,
                "created_by": p.created_by,
                "created_at": p.created_at.strftime("%Y-%m-%d %H:%M"),
            }
        )
    rows.sort(key=lambda r: r["created_at"])
    return rows


def _needs_attention(db: Session, master_db: Session, slug: str) -> list[dict]:
    """Items the Progress Matrix already flags as overdue or late, with the
    recovery suggestion it already computed — this report presents that work,
    it doesn't re-derive it."""
    matrix = progress_matrix.build_progress_matrix(
        slug=slug, db=db, master_db=master_db, entity_types=list(progress_matrix.ENTITY_MODELS)
    )
    rows = []
    for r in matrix["rows"]:
        if r["health"] not in ("overdue", "not_started_late", "late"):
            continue
        rows.append(
            {
                "module": r.get("item_type") or r["entity_type"],
                "entity_code": r["entity_code"] or f"#{r['entity_id']}",
                "entity_title": r["entity_title"],
                "owner": r["owner"],
                "state": r["health"].replace("_", " "),
                "plan_end": r["plan_end"] or "",
                "days_late": r["end_delay_days"] if r["end_delay_days"] is not None else r["start_delay_days"],
                "suggested_action": r["recovery"][0]["action"] if r["recovery"] else "",
            }
        )
    rows.sort(key=lambda r: (r["days_late"] is None, -(r["days_late"] or 0)))
    return rows


def generate(slug: str, target_date: date, db: Session, master_db: Session = None) -> Workbook:
    start_dt, end_dt = activity_feed.day_bounds(target_date)
    activity = activity_feed.fetch(db, start_dt, end_dt)
    status_rows = activity_feed.status_changes(activity)
    notes = _notes_created(db, start_dt, end_dt)
    attention = _needs_attention(db, master_db, slug) if master_db is not None else []

    wb = Workbook()

    # ---- 1. Summary ----
    ws = wb.active
    ws.title = "Summary"
    row = write_title(ws, 1, f"Daily Report — {slug} — {target_date.isoformat()}")
    row = write_note(ws, row, "For the delivery team: everything that moved today, in detail.")
    per_module = activity_feed.count_by(activity, "module")
    per_module_status = activity_feed.count_by(status_rows, "module")
    summary_rows = [
        {"module": m, "changes_today": c, "status_changes": per_module_status.get(m, 0)}
        for m, c in sorted(per_module.items(), key=lambda kv: -kv[1])
    ]

    row = write_section(ws, row, "Activity by module")
    if summary_rows:
        summary_rows.append(
            {"module": "TOTAL", "changes_today": len(activity), "status_changes": len(status_rows)}
        )
        row = write_table(ws, row, SUMMARY_COLUMNS, summary_rows)
    else:
        row = write_empty_notice(ws, row, SUMMARY_COLUMNS, "Nothing changed in any module on this date.")

    row = write_section(ws, row, "Other activity today")
    write_table(
        ws,
        row,
        ["item", "count"],
        [
            {"item": "Notes created", "count": len(notes)},
            {"item": "Items needing attention", "count": len(attention)},
        ],
    )

    # ---- 2. Activity Detail ----
    ws2 = wb.create_sheet("Activity Detail")
    row = write_title(ws2, 1, "Every change logged today")
    if activity:
        write_table(ws2, row, ACTIVITY_COLUMNS, activity)
    else:
        write_empty_notice(ws2, row, ACTIVITY_COLUMNS, "No changes were recorded in any module on this date.")

    # ---- 3. Status Changes ----
    ws3 = wb.create_sheet("Status Changes")
    row = write_title(ws3, 1, "Status changes only, grouped by module")
    if status_rows:
        for module in sorted({r["module"] for r in status_rows}):
            module_rows = [r for r in status_rows if r["module"] == module]
            row = write_section(ws3, row, f"{module} ({len(module_rows)})")
            row = write_table(ws3, row, STATUS_COLUMNS, module_rows)
    else:
        write_empty_notice(ws3, row, STATUS_COLUMNS, "Nothing changed status today.")

    # ---- 4. Notes Created ----
    ws4 = wb.create_sheet("Notes Created")
    row = write_title(ws4, 1, "Notes created today")
    if notes:
        write_table(ws4, row, NOTE_COLUMNS, notes)
    else:
        write_empty_notice(ws4, row, NOTE_COLUMNS, "No quick notes or note pages were created on this date.")

    # ---- 5. Needs Attention Today ----
    ws5 = wb.create_sheet("Needs Attention")
    row = write_title(ws5, 1, "Overdue and at-risk items")
    if attention:
        row = write_note(ws5, row, "From the Progress Matrix — suggestions are the ones it already computed.")
        write_table(ws5, row, ATTENTION_COLUMNS, attention)
    elif master_db is None:
        write_empty_notice(ws5, row, ATTENTION_COLUMNS, "Schedule analysis unavailable for this report run.")
    else:
        write_empty_notice(ws5, row, ATTENTION_COLUMNS, "Nothing is currently overdue or flagged at risk.")

    return wb

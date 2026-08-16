"""Weekly Report — audience: the delivery team plus steering.

Answers "is the shape of the week what we planned, and what is stuck". The
reader is one step back from the work, so this groups rather than lists:
activity by module and by person, board items by flow, documents by sign-off
stage. The line-by-line log lives in the Daily; the single-page view lives in
the Monthly.
"""

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from sqlalchemy.orm import Session

from .. import models, progress_matrix
from . import activity_feed
from .style import add_chart, write_empty_notice, write_note, write_section, write_table, write_title

PROGRESS_COLUMNS = [
    "phase",
    "functions_total",
    "functions_done",
    "functions_pct",
    "tasks_total",
    "tasks_done",
    "tasks_pct",
    "documents_total",
    "documents_confirmed",
    "documents_pct",
    "completed_this_week",
]
BOARD_COLUMNS = ["item_type", "severity", "opened", "closed", "still_open"]
SIGNOFF_COLUMNS = ["doc_code", "title", "status", "moved_to", "when", "by"]
MANDATORY_COLUMNS = ["doc_code", "doc_name", "phase_name", "current_status"]
CR_COLUMNS = ["cr_code", "title", "status", "requested_by", "target_date", "effort_md"]
AT_RISK_COLUMNS = ["module", "entity_code", "entity_title", "owner", "state", "days_late", "suggested_action"]
UTILIZATION_COLUMNS = ["resource", "avg_percent", "peak_percent", "weeks_over_100"]

FUNCTION_DONE_STATUSES = ("Confirmed", "Done")
TASK_DONE_STATUSES = ("Done",)
DOCUMENT_DONE_STATUSES = ("Confirmed",)
CLOSED_BOARD_STATUSES = ("Resolved", "Closed", "Done", "Promoted")


def _pct(done: int, total: int) -> str:
    return "-" if total == 0 else f"{round(done / total * 100)}%"


def generate(slug: str, week_start: date, db: Session, master_db: Session = None) -> Workbook:
    week_end = week_start + timedelta(days=7)
    start_dt, end_dt = activity_feed.range_bounds(week_start, week_end)
    activity = activity_feed.fetch(db, start_dt, end_dt)
    status_rows = activity_feed.status_changes(activity)
    days = [week_start + timedelta(days=i) for i in range(7)]

    wb = Workbook()

    # ---- 1. Progress Summary (with week-on-week movement) ----
    ws = wb.active
    ws.title = "Progress Summary"
    row = write_title(
        ws, 1, f"Weekly Report — {slug} — {week_start.isoformat()} to {(week_end - timedelta(days=1)).isoformat()}"
    )
    row = write_note(ws, row, "For the team and steering: grouped view of the week, not a line-by-line log.")

    functions = db.query(models.Function).all()
    tasks = db.query(models.Task).all()
    documents = db.query(models.Document).all()
    phases = sorted(
        {f.phase for f in functions if f.phase}
        | {t.phase for t in tasks if t.phase}
        | {d.phase for d in documents if d.phase},
        key=models.phase_sort_key,
    )

    # "Completed this week" is the delta the steering audience asks for. It is
    # counted from the log rather than a stored weekly snapshot, so it works
    # without a history table.
    done_values = set(FUNCTION_DONE_STATUSES) | set(TASK_DONE_STATUSES) | set(DOCUMENT_DONE_STATUSES)
    completed_by_phase: dict = {}
    for r in status_rows:
        if r["to"] not in done_values:
            continue
        source = activity_feed.ENTITY_SOURCES.get(r["entity_type"])
        if not source:
            continue
        obj = db.query(source[0]).filter(source[0].id == r["entity_id"]).first()
        phase = getattr(obj, "phase", None) if obj else None
        completed_by_phase[phase] = completed_by_phase.get(phase, 0) + 1

    progress_rows = []
    for phase in phases:
        f_in = [f for f in functions if f.phase == phase]
        t_in = [t for t in tasks if t.phase == phase]
        d_in = [d for d in documents if d.phase == phase]
        f_done = sum(1 for f in f_in if f.status in FUNCTION_DONE_STATUSES)
        t_done = sum(1 for t in t_in if t.status in TASK_DONE_STATUSES)
        d_done = sum(1 for d in d_in if d.status in DOCUMENT_DONE_STATUSES)
        progress_rows.append(
            {
                "phase": phase,
                "functions_total": len(f_in),
                "functions_done": f_done,
                "functions_pct": _pct(f_done, len(f_in)),
                "tasks_total": len(t_in),
                "tasks_done": t_done,
                "tasks_pct": _pct(t_done, len(t_in)),
                "documents_total": len(d_in),
                "documents_confirmed": d_done,
                "documents_pct": _pct(d_done, len(d_in)),
                "completed_this_week": f"+{completed_by_phase.get(phase, 0)}",
            }
        )

    row = write_section(ws, row, "% completion by phase (last column = completed during this week)")
    if progress_rows:
        header_row = row
        row = write_table(ws, row, PROGRESS_COLUMNS, progress_rows)
        # A numeric column for the chart to bind to — the % columns are text
        # ("83%") because that reads better in the table.
        pct_col = len(PROGRESS_COLUMNS) + 2
        ws.cell(row=header_row, column=pct_col, value="overall_pct")
        for i, r in enumerate(progress_rows):
            total = r["functions_total"] + r["tasks_total"] + r["documents_total"]
            done = r["functions_done"] + r["tasks_done"] + r["documents_confirmed"]
            ws.cell(row=header_row + 1 + i, column=pct_col, value=round(done / total * 100) if total else 0)
        chart = BarChart()
        chart.title = "% completion by phase"
        chart.y_axis.title = "%"
        add_chart(
            ws,
            chart,
            f"A{row + 1}",
            Reference(ws, min_col=pct_col, min_row=header_row, max_row=header_row + len(progress_rows)),
            Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(progress_rows)),
            len(progress_rows),
            skip_note_cell=f"A{row + 1}",
            reason="Completion chart omitted — needs at least 2 phases carrying items.",
        )
    else:
        write_empty_notice(ws, row, PROGRESS_COLUMNS, "No functions, tasks or documents carry a phase yet.")

    # ---- 2. Activity by Module × day ----
    ws2 = wb.create_sheet("Activity by Module")
    row = write_title(ws2, 1, "Activity by module, across the week")
    by_module_day = activity_feed.count_by_pair(activity, "module", "date")
    day_headers = [d.strftime("%a %d") for d in days]
    columns = ["module"] + day_headers + ["total"]
    if by_module_day:
        header_row = row
        rows = []
        for module, per_day in sorted(by_module_day.items()):
            entry = {"module": module}
            for d, header in zip(days, day_headers):
                entry[header] = per_day.get(d, 0)
            entry["total"] = sum(per_day.values())
            rows.append(entry)
        row = write_table(ws2, header_row, columns, rows)
        chart = BarChart()
        chart.title = "Changes per day, by module"
        chart.y_axis.title = "changes"
        add_chart(
            ws2,
            chart,
            f"A{row + 1}",
            Reference(ws2, min_col=2, max_col=1 + len(day_headers), min_row=header_row, max_row=header_row + len(rows)),
            Reference(ws2, min_col=1, min_row=header_row + 1, max_row=header_row + len(rows)),
            len(rows),
            skip_note_cell=f"A{row + 1}",
            reason="Activity chart omitted — needs at least 2 modules with activity this week.",
        )
    else:
        write_empty_notice(ws2, row, columns, "No changes were recorded in any module this week.")

    # ---- 3. Activity by Person ----
    ws3 = wb.create_sheet("Activity by Person")
    row = write_title(ws3, 1, "Who did what, by module")
    named = [r for r in activity if r["by"]]
    by_person = activity_feed.count_by_pair(named, "by", "module")
    modules = sorted({r["module"] for r in named})
    if by_person:
        rows = []
        for person, per_module in sorted(by_person.items()):
            entry = {"person": person, **{m: per_module.get(m, 0) for m in modules}}
            entry["total"] = sum(per_module.values())
            rows.append(entry)
        write_table(ws3, row, ["person"] + modules + ["total"], rows)
    else:
        write_empty_notice(
            ws3,
            row,
            ["person", "total"],
            "No changes this week recorded who made them — attribution depends on the caller supplying a name.",
        )

    # ---- 4. Board Item Flow ----
    ws4 = wb.create_sheet("Board Item Flow")
    row = write_title(ws4, 1, "Issues, incidents and backlog — opened vs closed this week")
    board_items = db.query(models.BoardItem).all()
    board_status = [r for r in status_rows if r["entity_type"] == "board_item"]
    board_rows = []
    for item_type in ("issue", "incident", "backlog"):
        for severity in ("Critical", "High", "Medium", "Low", None):
            group = [b for b in board_items if b.item_type == item_type and b.severity == severity]
            if not group:
                continue
            group_ids = {b.id for b in group}
            board_rows.append(
                {
                    "item_type": item_type,
                    "severity": severity or "(none)",
                    "opened": sum(1 for b in group if b.created_at and start_dt <= b.created_at < end_dt),
                    "closed": sum(
                        1 for r in board_status if r["entity_id"] in group_ids and r["to"] in CLOSED_BOARD_STATUSES
                    ),
                    "still_open": sum(1 for b in group if b.status not in CLOSED_BOARD_STATUSES),
                }
            )
    if board_rows:
        header_row = row
        row = write_table(ws4, header_row, BOARD_COLUMNS, board_rows)
        chart = BarChart()
        chart.title = "Board items opened vs closed"
        add_chart(
            ws4,
            chart,
            f"A{row + 1}",
            Reference(ws4, min_col=3, max_col=4, min_row=header_row, max_row=header_row + len(board_rows)),
            Reference(ws4, min_col=2, min_row=header_row + 1, max_row=header_row + len(board_rows)),
            len(board_rows),
            skip_note_cell=f"A{row + 1}",
            reason="Board flow chart omitted — needs at least 2 severity groups to compare.",
        )
    else:
        write_empty_notice(ws4, row, BOARD_COLUMNS, "No issues, incidents or backlog items exist in this project yet.")

    # ---- 5. Document Sign-off Progress ----
    ws5 = wb.create_sheet("Document Sign-off")
    row = write_title(ws5, 1, "Documents that moved through sign-off this week")
    doc_moves = [r for r in status_rows if r["entity_type"] == "document"]
    if doc_moves:
        row = write_table(
            ws5,
            row,
            SIGNOFF_COLUMNS,
            [
                {
                    "doc_code": r["entity_code"],
                    "title": r["entity_title"],
                    "status": r["from"],
                    "moved_to": r["to"],
                    "when": r["datetime"],
                    "by": r["by"],
                }
                for r in doc_moves
            ],
        )
    else:
        row = write_empty_notice(ws5, row, SIGNOFF_COLUMNS, "No document changed status this week.")

    row = write_section(ws5, row, "Mandatory documents still outstanding")
    pending = _mandatory_pending(db, master_db, slug) if master_db is not None else []
    if pending:
        write_table(ws5, row, MANDATORY_COLUMNS, pending)
    else:
        write_empty_notice(
            ws5,
            row,
            MANDATORY_COLUMNS,
            "No outstanding mandatory documents — or this project has no category set, so none are required.",
        )

    # ---- 6. Change Requests ----
    ws6 = wb.create_sheet("Change Requests")
    row = write_title(ws6, 1, "Change requests raised or moved this week")
    changed_cr_ids = {r["entity_id"] for r in status_rows if r["entity_type"] == "change_request"}
    all_crs = db.query(models.ChangeRequest).all()
    crs = [
        c
        for c in all_crs
        if (c.created_at and start_dt <= c.created_at < end_dt) or c.id in changed_cr_ids
    ]
    if crs:
        estimates: dict = {}
        for e in (
            db.query(models.EffortEstimate)
            .filter(models.EffortEstimate.linked_entity_type == "change_request")
            .all()
        ):
            estimates[e.linked_entity_id] = estimates.get(e.linked_entity_id, 0) + (e.calculated_man_days or 0)
        row = write_table(
            ws6,
            row,
            CR_COLUMNS,
            [
                {
                    "cr_code": c.cr_code,
                    "title": c.title,
                    "status": c.status,
                    "requested_by": c.requested_by,
                    "target_date": c.target_date.isoformat() if c.target_date else "",
                    "effort_md": round(estimates.get(c.id, 0), 2) or "",
                }
                for c in crs
            ],
        )
        write_note(ws6, row, f"Total effort across these change requests: {sum(estimates.get(c.id, 0) for c in crs):.2f} MD")
    else:
        write_empty_notice(ws6, row, CR_COLUMNS, "No change requests were raised or changed status this week.")

    # ---- 7. At Risk ----
    ws7 = wb.create_sheet("At Risk")
    row = write_title(ws7, 1, "Items the Progress Matrix flags as late or overdue")
    at_risk = _at_risk(db, master_db, slug) if master_db is not None else []
    if at_risk:
        write_table(ws7, row, AT_RISK_COLUMNS, at_risk)
    elif master_db is None:
        write_empty_notice(ws7, row, AT_RISK_COLUMNS, "Schedule analysis unavailable for this report run.")
    else:
        write_empty_notice(ws7, row, AT_RISK_COLUMNS, "Nothing is currently flagged late or overdue.")

    # ---- 8. Resource Utilization (internal) ----
    ws8 = wb.create_sheet("Resource Utilization")
    row = write_title(ws8, 1, "Resource utilization this week")
    row = write_note(ws8, row, "Internal only — this sheet is never included in client-facing output.")
    util = _utilization(master_db, week_start, week_end) if master_db is not None else []
    if util:
        write_table(ws8, row, UTILIZATION_COLUMNS, util)
    else:
        write_empty_notice(ws8, row, UTILIZATION_COLUMNS, "No resources are allocated over this week.")

    return wb


def _mandatory_pending(db: Session, master_db: Session, slug: str) -> list[dict]:
    from ..routers.projects import MANDATORY_COLUMN_BY_CATEGORY

    project = master_db.query(models.Project).filter(models.Project.slug == slug).first()
    column = MANDATORY_COLUMN_BY_CATEGORY.get(project.project_category) if project else None
    if not column:
        return []
    docs = {d.doc_code: d for d in db.query(models.Document).all()}
    rows = []
    for t in (
        master_db.query(models.DocumentTemplate)
        .filter(getattr(models.DocumentTemplate, column) == "M")
        .order_by(models.DocumentTemplate.doc_code)
        .all()
    ):
        doc = docs.get(str(t.doc_code))
        if doc and doc.status == "Confirmed":
            continue
        rows.append(
            {
                "doc_code": t.doc_code,
                "doc_name": t.doc_name,
                "phase_name": t.phase_name,
                "current_status": doc.status if doc else "(not created)",
            }
        )
    return rows


def _at_risk(db: Session, master_db: Session, slug: str) -> list[dict]:
    matrix = progress_matrix.build_progress_matrix(
        slug=slug, db=db, master_db=master_db, entity_types=list(progress_matrix.ENTITY_MODELS)
    )
    rows = [
        {
            "module": r.get("item_type") or r["entity_type"],
            "entity_code": r["entity_code"] or f"#{r['entity_id']}",
            "entity_title": r["entity_title"],
            "owner": r["owner"],
            "state": r["health"].replace("_", " "),
            "days_late": r["end_delay_days"] if r["end_delay_days"] is not None else r["start_delay_days"],
            "suggested_action": r["recovery"][0]["action"] if r["recovery"] else "",
        }
        for r in matrix["rows"]
        if r["health"] in ("overdue", "not_started_late", "late")
    ]
    rows.sort(key=lambda r: (r["days_late"] is None, -(r["days_late"] or 0)))
    return rows


def _utilization(master_db: Session, week_start: date, week_end: date) -> list[dict]:
    from ..resource_utils import compute_utilization

    rows = []
    for r in compute_utilization(master_db, week_start, week_end - timedelta(days=1)):
        percents = [w.get("total_percent") or 0 for w in (r.get("weeks") or [])]
        if not percents or not any(percents):
            continue
        rows.append(
            {
                "resource": r["resource_name"],
                "avg_percent": round(sum(percents) / len(percents)),
                "peak_percent": max(percents),
                "weeks_over_100": sum(1 for p in percents if p > 100),
            }
        )
    return rows

"""Impact Analysis document (xlsx) for a Change Request.

Laid out to follow the customer's own `Impact Analysis` sheet in
reference/Impact_Analysis_TBOS_*.xlsx, section for section:

    Header            Phase / Block Name / Document Name | Project / Function
                      Name / Title
    General Information   Incident No, Target Function, Target Date
    Efforts Estimation (Man-day)   total + DR / DN&PU / IFT-BCT
    Target Function Description
    Function Impact   the impacted function table
    Database Impact   present in the real workbook, absent from the kickoff
                      spec — included so the generated file is a drop-in
                      replacement for the hand-made one

Rounding note: the real workbook computes the headline total as
ROUND(SUM(MM), 1) * 20, i.e. it rounds man-months to one decimal BEFORE
converting to man-days, while the per-phase cells round only at the end. In
the reference file that makes the header read "12 MD" while the phases add up
to 12.9. That behaviour is reproduced here (so a generated document matches
one produced by hand) and the unrounded figure is written alongside it rather
than silently dropped.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from .. import models
from .style import BODY_FONT, SECTION_FONT, TITLE_FONT, write_section, write_table

DELIVERY_MODE_LABELS = {"human": "HUMAN", "human_in_loop": "HUMAN-in-LOOP"}

IMPACT_COLUMNS = ["Function Name", "Impact Type", "Impact Description"]
DATABASE_COLUMNS = ["Table Name", "In/Out", "Field/Table Impact", "Impact Type", "Impact Description"]

LABEL_FONT = Font(name="Calibri", bold=True)


def _label(ws, row: int, label: str, value, value_col: int = 3) -> int:
    ws.cell(row=row, column=2, value=label).font = LABEL_FONT
    cell = ws.cell(row=row, column=value_col, value=value)
    cell.font = BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _excel_round(value: float, digits: int = 1) -> float:
    """Excel ROUND — half away from zero, unlike Python's banker's rounding."""
    if value is None:
        return None
    factor = 10 ** digits
    scaled = value * factor
    return (int(scaled + 0.5) if scaled >= 0 else -int(-scaled + 0.5)) / factor


def generate(slug: str, cr: models.ChangeRequest, impact: dict, db: Session, master_db: Session) -> Workbook:
    project = master_db.query(models.Project).filter(models.Project.slug == slug).first()
    impacts = (
        db.query(models.ChangeRequestImpact)
        .filter(models.ChangeRequestImpact.change_request_id == cr.id)
        .order_by(models.ChangeRequestImpact.id)
        .all()
    )
    function_ids = [i.linked_function_id for i in impacts if i.linked_function_id]
    functions = {}
    if function_ids:
        functions = {
            f.id: f for f in db.query(models.Function).filter(models.Function.id.in_(function_ids)).all()
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Analysis"

    config = db.query(models.EffortEstimateConfig).order_by(models.EffortEstimateConfig.id).first()
    # Off by default: a client document states the man-days and the price, not
    # how the work is produced. Only a client who has explicitly asked to see
    # the delivery model gets this section.
    show_delivery_mode = bool(config.show_delivery_mode_in_client_docs) if config else False

    effort = impact.get("effort") or {}
    phases = {k: effort.get(k) for k in ("dr", "dnpu", "iftbct")}
    total_mm = None
    if effort.get("total_md") is not None:
        working_days = (config.working_days_per_month if config else None) or 20
        total_mm = effort["total_md"] / working_days

    estimates = (
        db.query(models.EffortEstimate)
        .filter(
            models.EffortEstimate.linked_entity_type == "change_request",
            models.EffortEstimate.linked_entity_id == cr.id,
        )
        .all()
    )

    row = 1
    ws.cell(row=row, column=2, value="Impact Analysis").font = TITLE_FONT
    row += 2

    # ---- header block ----
    row = _label(ws, row, "Phase", cr.status)
    row = _label(ws, row, "Block Name", (functions.get(function_ids[0]).module if function_ids and functions.get(function_ids[0]) else None) or "Common")
    row = _label(ws, row, "Document Name", "Impact Analysis")
    row = _label(ws, row, "Project Name", project.name if project else slug)
    row = _label(ws, row, "Function Name", impacts[0].function_name if impacts else cr.title)
    row = _label(ws, row, "Title", cr.title)
    row = _label(ws, row, "Created by", cr.requested_by)
    row = _label(ws, row, "Created date", cr.created_at.date().isoformat() if cr.created_at else None)
    row = _label(ws, row, "Updated date", cr.updated_at.date().isoformat() if cr.updated_at else None)
    row += 1

    # ---- general information ----
    row = write_section(ws, row, "General Information")
    row = _label(ws, row, "Incident No :", cr.cr_code or "(no CR code assigned yet)")
    row = _label(ws, row, "Target Function :", impacts[0].function_name if impacts else "-")
    row = _label(ws, row, "Target Date :", cr.target_date.isoformat() if cr.target_date else "TBD")
    row += 1

    # ---- efforts estimation ----
    row = write_section(ws, row, "Efforts Estimation (Man-day)")
    if effort.get("total_md") is None:
        row = _label(ws, row, "Total :", "no effort estimate linked to this change request")
    else:
        headline = _excel_round(_excel_round(total_mm, 1) * 20, 1)
        row = _label(ws, row, "Total :", f"{headline:g} MD")
        # The unrounded figure, so nobody has to reverse-engineer the gap
        # between the headline and the phase columns.
        row = _label(ws, row, "Total (unrounded) :", f"{effort['total_md']:.4f} MD")
        ws.cell(row=row, column=3, value="DR").font = LABEL_FONT
        ws.cell(row=row, column=4, value="DN&PU").font = LABEL_FONT
        ws.cell(row=row, column=5, value="IFT/BCT").font = LABEL_FONT
        row += 1
        ws.cell(row=row, column=2, value="By phase :").font = LABEL_FONT
        for col, key in ((3, "dr"), (4, "dnpu"), (5, "iftbct")):
            value = phases.get(key)
            ws.cell(row=row, column=col, value=f"{_excel_round(value, 1):g} MD" if value is not None else "-").font = BODY_FONT
        row += 2

        # Delivery model — suppressed unless the project has opted in.
        if show_delivery_mode and estimates:
            row = write_section(ws, row, "Delivery Model")
            modes = sorted({(e.delivery_mode or "human") for e in estimates})
            row = _label(ws, row, "Mode :", ", ".join(DELIVERY_MODE_LABELS.get(m, m) for m in modes))
            human_total = sum(e.man_days_human or 0 for e in estimates)
            if human_total:
                row = _label(ws, row, "Fully-human equivalent :", f"{_excel_round(human_total, 1):g} MD")
                saved = human_total - (effort.get("total_md") or 0)
                if saved > 0:
                    row = _label(
                        ws,
                        row,
                        "Reduction :",
                        f"{_excel_round(saved, 1):g} MD ({saved / human_total * 100:.0f}%)",
                    )
            row += 1

    # ---- target function description ----
    row = write_section(ws, row, "Target Function Description")
    ws.cell(row=row, column=2, value=cr.description or "-").font = BODY_FONT
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    row += 2

    # ---- function impact ----
    row = write_section(ws, row, "Function Impact")
    impact_rows = []
    for i in impacts:
        function = functions.get(i.linked_function_id) if i.linked_function_id else None
        impact_rows.append(
            {
                "Function Name": i.function_name or (function.name if function else "(new function)"),
                "Impact Type": i.impact_type,
                "Impact Description": i.note or "",
            }
        )
    row = write_table(ws, row, IMPACT_COLUMNS, impact_rows)

    # ---- database impact ----
    # No structured source for this in the app yet, so the section is emitted
    # with its headers only rather than being left out — the person finishing
    # the document fills it in, exactly as they do today.
    row = write_section(ws, row, "Database Impact")
    row = write_table(ws, row, DATABASE_COLUMNS, [])

    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 42
    for col in ("D", "E", "F"):
        ws.column_dimensions[col].width = 18
    ws.cell(row=row, column=2, value="Generated by PM-Again from the linked effort estimates.").font = SECTION_FONT

    return wb

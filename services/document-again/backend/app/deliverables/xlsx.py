"""R17 — Reusable XLSX deliverable renderer.

Formatting is implemented once as primitives (cover, document control,
revision history, index, register, sign-off, source reference, print/brand
application) and reused by every workbook. Content is sourced from the
project's authoritative truth — nothing is fabricated; missing data is shown
as MISSING / TBD / NOT YET RECORDED.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import models as m
from . import service as dsvc
from .layouts import get_brand, layout_registry
from .standards import get_standard

# Brand palette defaults (overridden by brand profile)
NAVY = "1F4E78"
PALE_BLUE = "D9EAF7"
PALE_GRAY = "F3F4F6"
WHITE = "FFFFFF"
BORDER = "D1D5DB"


# ────────────────────────────────────────────────────────────────────────────
# Context collection (authoritative project truth)
# ────────────────────────────────────────────────────────────────────────────
def build_context(db: Session, project: m.Project) -> dict:
    reqs = db.execute(select(m.Requirement).where(m.Requirement.project_id == project.id)).scalars().all()
    assumptions = db.execute(select(m.Assumption).where(m.Assumption.project_id == project.id)).scalars().all()
    decisions = db.execute(select(m.Decision).where(m.Decision.project_id == project.id)).scalars().all()
    clarifications = db.execute(select(m.Clarification).where(m.Clarification.project_id == project.id)).scalars().all()
    crs = db.execute(select(m.ChangeRequest).where(m.ChangeRequest.project_id == project.id)).scalars().all()
    baselines = db.execute(select(m.Baseline).where(m.Baseline.project_id == project.id)).scalars().all()
    traces = db.execute(select(m.TraceLink).where(m.TraceLink.project_id == project.id)).scalars().all()

    arch_diagrams = db.execute(
        select(m.ArchitectureDiagram).where(m.ArchitectureDiagram.project_id == project.id)
    ).scalars().all()
    diagrams = []
    for d in arch_diagrams:
        nodes = db.execute(select(m.ArchitectureNode).where(m.ArchitectureNode.diagram_id == d.id)).scalars().all()
        diagrams.append({"name": d.name, "nodes": [n for n in nodes]})

    flows = db.execute(select(m.ProcessFlow).where(m.ProcessFlow.project_id == project.id)).scalars().all()
    flow_list = []
    for f in flows:
        steps = db.execute(select(m.ProcessStep).where(m.ProcessStep.flow_id == f.id)).scalars().all()
        flow_list.append({"name": f.name, "steps": [s for s in steps]})

    matrix = dsvc.get_matrix(db, project)["rows"]

    return {
        "project": {
            "id": project.id, "key": project.key, "name": project.name,
            "description": project.description, "lifecycle_state": project.lifecycle_state,
            "created_by": project.created_by,
        },
        "profile": dsvc.get_profile(db, project),
        "matrix": matrix,
        "summary": dsvc.completeness(matrix),
        "requirements": [
            {"code": r.code, "title": r.title, "status": r.status.value if hasattr(r.status, "value") else str(r.status),
             "priority": r.priority, "source": r.source_type}
            for r in reqs
        ],
        "assumptions": [
            {"id": a.semantic_id, "content": a.content, "status": a.status} for a in assumptions
        ],
        "decisions": [
            {"id": d.semantic_id, "title": d.title, "content": d.content} for d in decisions
        ],
        "clarifications": [
            {"question": c.question, "answer": c.answer, "resolved": c.resolved} for c in clarifications
        ],
        "change_requests": [
            {"code": cr.code, "title": cr.title or cr.requested_change,
             "requested_change": cr.requested_change,
             "status": cr.status.value if hasattr(cr.status, "value") else str(cr.status)}
            for cr in crs
        ],
        "baselines": [{"name": b.name, "created_at": b.created_at.isoformat()} for b in baselines],
        "trace_links": [
            {"source": t.source_semantic_id, "target": t.target_semantic_id,
             "relation": t.relation_type.value if hasattr(t.relation_type, "value") else str(t.relation_type)}
            for t in traces
        ],
        "architecture": diagrams,
        "flows": flow_list,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


# ────────────────────────────────────────────────────────────────────────────
# Renderer primitives
# ────────────────────────────────────────────────────────────────────────────
class Xlsx:
    def __init__(self, brand_name: str = "GEA_STANDARD"):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        self.openpyxl = openpyxl
        brand = get_brand(brand_name)
        c = brand["colors"]
        self.primary = c["primary"]
        self.secondary = c["secondary"]
        self.neutral = c["neutral"]
        self.border_color = c["border"]
        self.accent = c["accent"]
        self.classification = brand["classification_default"]
        self.company = brand["company_name"] or "Global Edge"

        self.Alignment = Alignment
        self.Border = Border
        self.Font = Font
        self.PatternFill = PatternFill
        self.Side = Side
        self.thin = Side(style="thin", color=self.border_color)

    # -- low-level helpers ---------------------------------------------------
    def new_workbook(self):
        wb = self.openpyxl.Workbook()
        wb.remove(wb.active)
        return wb

    def setup_page(self, sheet, landscape=False, repeat=None):
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape" if landscape else "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = sheet.page_margins.right = 0.3
        sheet.page_setup.paperSize = 9  # A4
        if repeat:
            sheet.print_title_rows = repeat

    def header_row(self, sheet, row):
        for cell in sheet[row]:
            if cell.value is not None:
                cell.fill = self.PatternFill("solid", fgColor=self.primary)
                cell.font = self.Font(bold=True, color=WHITE)
                cell.alignment = self.Alignment(wrap_text=True, vertical="center")
                cell.border = self.Border(bottom=self.thin)
        sheet.row_dimensions[row].height = 26

    def footer_text(self, project_key, doc_id, version, classification):
        return f"{project_key} | {doc_id} | Version {version} | {classification} | Page &P of &N"

    def _apply_footer(self, sheet, project_key, doc_id, version):
        sheet.oddFooter.center.text = self.footer_text(project_key, doc_id, version, self.classification)
        sheet.oddFooter.center.size = 8

    # -- cover ---------------------------------------------------------------
    def render_cover(self, wb, doc: dict):
        ws = wb.create_sheet("00_Cover")
        self.setup_page(ws)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        for i in range(1, 25):
            ws.cell(row=i, column=1).fill = self.PatternFill("solid", fgColor=WHITE)
            ws.cell(row=i, column=2).fill = self.PatternFill("solid", fgColor=WHITE)

        # header band
        ws.merge_cells("A1:B1")
        c = ws["A1"]
        c.value = self.company
        c.font = self.Font(bold=True, size=20, color=WHITE)
        c.fill = self.PatternFill("solid", fgColor=self.primary)
        c.alignment = self.Alignment(vertical="center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:B2")
        c = ws["A2"]
        c.value = doc.get("customer") or "Customer"
        c.font = self.Font(bold=True, size=13, color=self.accent)
        c.fill = self.PatternFill("solid", fgColor=self.neutral)
        ws.row_dimensions[2].height = 24

        ws.merge_cells("A3:B4")
        c = ws["A3"]
        c.value = doc.get("title", "Project Deliverable")
        c.font = self.Font(bold=True, size=16, color="1F2937")
        c.alignment = self.Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 26
        ws.row_dimensions[4].height = 26

        fields = [
            ("Project", doc.get("project")),
            ("Project Code", doc.get("project_code")),
            ("Document ID", doc.get("document_id")),
            ("Template ID", doc.get("template_id")),
            ("Template Version", doc.get("template_version")),
            ("Document Version", doc.get("version")),
            ("Revision", doc.get("revision")),
            ("Status", doc.get("status")),
            ("Classification", doc.get("classification") or self.classification),
            ("Prepared By", doc.get("prepared_by")),
            ("Reviewed By", doc.get("reviewed_by")),
            ("Approved By", doc.get("approved_by")),
            ("Effective Date", doc.get("effective_date")),
        ]
        r = 6
        for label, value in fields:
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = self.Font(bold=True, color="374151")
            lc.fill = self.PatternFill("solid", fgColor=self.neutral)
            lc.border = self.Border(bottom=self.Side(style="hair", color=self.border_color))
            vc = ws.cell(row=r, column=2, value=(value if value not in (None, "") else "—"))
            vc.alignment = self.Alignment(wrap_text=True, vertical="top")
            vc.border = self.Border(bottom=self.Side(style="hair", color=self.border_color))
            r += 1
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- document control ----------------------------------------------------
    def render_document_control(self, wb, doc: dict, related: list | None = None):
        ws = wb.create_sheet("01_Document_Control")
        self.setup_page(ws)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 46
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.append(["Field", "Value", None, None])
        self.header_row(ws, 1)
        ws.merge_cells("A1:D1")
        fields = [
            ("Document ID", doc.get("document_id")),
            ("Document Title", doc.get("title")),
            ("Project", doc.get("project")),
            ("Project Code", doc.get("project_code")),
            ("Customer", doc.get("customer")),
            ("Template ID", doc.get("template_id")),
            ("Template Version", doc.get("template_version")),
            ("Document Version", doc.get("version")),
            ("Revision", doc.get("revision")),
            ("Status", doc.get("status")),
            ("Classification", doc.get("classification") or self.classification),
            ("Owner", doc.get("owner")),
            ("Prepared By", doc.get("prepared_by")),
            ("Reviewed By", doc.get("reviewed_by")),
            ("Approved By", doc.get("approved_by")),
            ("Effective Date", doc.get("effective_date")),
        ]
        r = 2
        for label, value in fields:
            ws.cell(row=r, column=1, value=label).font = self.Font(bold=True)
            ws.cell(row=r, column=1).fill = self.PatternFill("solid", fgColor=self.neutral)
            ws.cell(row=r, column=2, value=(value if value not in (None, "") else "—"))
            r += 1
        # related documents
        r += 1
        ws.cell(row=r, column=1, value="Related Documents").font = self.Font(bold=True)
        r += 1
        ws.append(["Document ID", "Title", "Version", "Relationship"])
        self.header_row(ws, r)
        hdr = r
        for rel in (related or []):
            r += 1
            ws.append([rel.get("document_id"), rel.get("title"), rel.get("version"), rel.get("relationship")])
        if not related:
            r += 1
            ws.cell(row=r, column=1, value="— none recorded —")
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- revision history ----------------------------------------------------
    def render_revision_history(self, wb, doc: dict, revisions: list | None = None):
        ws = wb.create_sheet("02_Revision_History")
        self.setup_page(ws, landscape=True, repeat="1:1")
        ws.freeze_panes = "A2"
        headers = ["Revision", "Date", "Description of Change", "Prepared By", "Reviewed By", "Approval Status"]
        ws.append(headers)
        self.header_row(ws, 1)
        ws.auto_filter.ref = f"A1:F1"
        widths = (12, 14, 52, 18, 18, 18)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[self.openpyxl.utils.get_column_letter(i)].width = w
        for rv in (revisions or [{"revision": doc.get("revision", "0.1"), "date": doc.get("effective_date"),
                                  "description": "Initial draft", "prepared": doc.get("prepared_by"),
                                  "reviewed": doc.get("reviewed_by"), "status": doc.get("status")}]):
            ws.append([rv.get("revision"), rv.get("date"), rv.get("description"),
                       rv.get("prepared"), rv.get("reviewed"), rv.get("status")])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = self.Alignment(wrap_text=True, vertical="top")
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- sheet index ---------------------------------------------------------
    def render_index(self, wb, doc: dict, entries: list[dict]):
        ws = wb.create_sheet("03_Sheet_Index")
        self.setup_page(ws, repeat="1:1")
        ws.freeze_panes = "A2"
        headers = ["No.", "Sheet", "Description", "Applicability", "Status", "Owner"]
        ws.append(headers)
        self.header_row(ws, 1)
        for i, e in enumerate(entries, start=1):
            ws.append([i, e.get("sheet"), e.get("description"), e.get("applicability", ""),
                       e.get("status", ""), e.get("owner", "")])
            cell = ws.cell(row=i + 1, column=2)
            cell.hyperlink = f"#'{e.get('sheet')}'!A1"
            cell.font = self.Font(color=self.accent, underline="single")
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- generic register ----------------------------------------------------
    def render_register(self, wb, title: str, headers: list[str], rows: list[list],
                        doc: dict, landscape=True, widths=None):
        sheet_name = title[:31]
        ws = wb.create_sheet(sheet_name)
        self.setup_page(ws, landscape=landscape, repeat="1:3")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(row=1, column=1, value=title)
        c.font = self.Font(bold=True, size=14, color=WHITE)
        c.fill = self.PatternFill("solid", fgColor=self.primary)
        ws.row_dimensions[1].height = 26
        ws.cell(row=2, column=1, value=f"{doc.get('project')} · {doc.get('document_id')} · {doc.get('status', 'DRAFT')}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1).font = self.Font(size=9, color="6B7280")
        ws.cell(row=3, column=1, value="")
        for j, h in enumerate(headers, start=1):
            ws.cell(row=4, column=j, value=h)
        self.header_row(ws, 4)
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{self.openpyxl.utils.get_column_letter(len(headers))}4"
        if widths:
            for j, w in enumerate(widths, start=1):
                ws.column_dimensions[self.openpyxl.utils.get_column_letter(j)].width = w
        if rows:
            for row in rows:
                ws.append(row)
        else:
            ws.append(["TBD — NOT YET RECORDED"] + [""] * (len(headers) - 1))
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                cell.alignment = self.Alignment(wrap_text=True, vertical="top")
                cell.border = self.Border(bottom=self.Side(style="hair", color=self.border_color))
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- sign-off ------------------------------------------------------------
    def render_signoff(self, wb, doc: dict, signoffs: list | None = None):
        ws = wb.create_sheet("90_Review_Signoff")
        self.setup_page(ws)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 20
        ws.append(["Role", "Name", "Signature", "Date"])
        self.header_row(ws, 1)
        rows = signoffs or [
            {"role": "Prepared By", "name": doc.get("prepared_by") or "—"},
            {"role": "Reviewed By", "name": doc.get("reviewed_by") or "—"},
            {"role": "Approved By", "name": doc.get("approved_by") or "—"},
        ]
        for s in rows:
            ws.append([s.get("role"), s.get("name"), "", s.get("date") or ""])
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws

    # -- source reference ----------------------------------------------------
    def render_source_reference(self, wb, doc: dict, sources: list | None = None):
        ws = wb.create_sheet("99_Source_Reference")
        self.setup_page(ws, landscape=True, repeat="1:1")
        ws.freeze_panes = "A2"
        headers = ["Authority", "Object Type", "Object ID", "Version", "Retrieved At"]
        ws.append(headers)
        self.header_row(ws, 1)
        for s in (sources or []):
            ws.append([s.get("authority"), s.get("object_type"), s.get("object_id"),
                       s.get("version"), s.get("retrieved_at")])
        if not sources:
            ws.append(["DOCUMENT_AGAIN", "Project", doc.get("project_code"), "—", doc.get("generated_at")])
        self._apply_footer(ws, doc.get("project_code") or "", doc.get("document_id") or "", doc.get("version") or "—")
        return ws


# ────────────────────────────────────────────────────────────────────────────
# Document descriptor + workbook builders
# ────────────────────────────────────────────────────────────────────────────
def _base_doc(ctx: dict, mode: str, title: str) -> dict:
    p = ctx["project"]
    return {
        "title": title,
        "project": p["name"],
        "project_code": p["key"],
        "document_id": f"{p['key']}-{mode}",
        "template_id": f"LAYOUT-{mode}-001",
        "template_version": "1.0",
        "version": "0.1",
        "revision": "0.1",
        "status": "DRAFT — NOT APPROVED",
        "classification": None,
        "prepared_by": p.get("created_by") or "—",
        "reviewed_by": "—",
        "approved_by": "—",
        "effective_date": ctx.get("generated_at", "")[:10],
        "generated_at": ctx.get("generated_at"),
    }


def _index_entries(sheet_names: list[str]) -> list[dict]:
    out = []
    for s in sheet_names:
        out.append({"sheet": s, "description": "", "applicability": "", "status": "", "owner": ""})
    return out


def build_master_workbook(ctx: dict, brand: str = "GEA_STANDARD") -> bytes:
    x = Xlsx(brand)
    wb = x.new_workbook()
    doc = _base_doc(ctx, "PROJECT_MASTER", f"{ctx['project']['name']} — Project Master")
    profile = ctx.get("profile", {})

    x.render_cover(wb, doc)
    x.render_document_control(wb, doc)
    x.render_revision_history(wb, doc)
    sheet_names = ["00_Cover", "01_Document_Control", "02_Revision_History", "03_Sheet_Index",
                   "01_Project_Context", "02_Deliverable_Matrix", "03_Milestones", "04_Stakeholders",
                   "10_Requirement_Register", "11_Clarification_Register", "12_Assumption_Register",
                   "13_Decision_Register", "14_Dependency_Register", "20_RAID", "21_Action_Issue",
                   "30_Traceability", "31_Change_Request", "40_Acceptance", "41_Handover",
                   "90_Review_Signoff", "99_Source_Reference"]
    x.render_index(wb, doc, _index_entries(sheet_names))

    # 01 Project Context
    ws = wb.create_sheet("01_Project_Context")
    x.setup_page(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.merge_cells("A1:B1")
    c = ws["A1"]; c.value = "Project Context"; c.font = x.Font(bold=True, size=14, color=WHITE); c.fill = x.PatternFill("solid", fgColor=x.primary)
    fields = [
        ("Project", ctx["project"]["name"]),
        ("Project Code", ctx["project"]["key"]),
        ("Primary Project Type", profile.get("primary_type") or "TBD"),
        ("Workstreams", ", ".join(profile.get("workstreams") or []) or "TBD"),
        ("Production Impact", profile.get("attributes", {}).get("production_impact") or "TBD"),
        ("Regulated", "Yes" if profile.get("attributes", {}).get("regulated") else "No"),
        ("Data Migration", "Yes" if profile.get("attributes", {}).get("data_migration") else "No"),
        ("Description", ctx["project"].get("description") or "—"),
    ]
    r = 3
    for label, val in fields:
        ws.cell(row=r, column=1, value=label).font = x.Font(bold=True)
        ws.cell(row=r, column=1).fill = x.PatternFill("solid", fgColor=x.neutral)
        ws.cell(row=r, column=2, value=str(val)).alignment = x.Alignment(wrap_text=True, vertical="top")
        r += 1

    # 02 Deliverable Matrix
    x.render_register(wb, "Deliverable Matrix",
                      ["Deliverable ID", "Name", "Domain", "Category", "Applicability", "Status", "Owner", "Version"],
                      [[r["document_id"], r["name"], r["domain"], r["category"], r["applicability"],
                        r["lifecycle_status"], r["owner"] or "—", r["version"] or "—"] for r in ctx["matrix"]],
                      doc, widths=(22, 34, 16, 14, 14, 16, 16, 10))

    # registers
    x.render_register(wb, "10_Requirement_Register",
                      ["Code", "Title", "Status", "Priority", "Source"],
                      [[r["code"], r["title"], r["status"], r["priority"], r["source"]] for r in ctx.get("requirements", [])],
                      doc, widths=(14, 52, 12, 10, 18))
    x.render_register(wb, "11_Clarification_Register",
                      ["Question", "Answer", "Resolved"],
                      [[c["question"], c["answer"] or "—", "Yes" if c["resolved"] else "No"] for c in ctx.get("clarifications", [])],
                      doc, widths=(46, 46, 10))
    x.render_register(wb, "12_Assumption_Register",
                      ["ID", "Assumption", "Status"],
                      [[a["id"], a["content"], a["status"]] for a in ctx.get("assumptions", [])],
                      doc, widths=(14, 76, 12))
    x.render_register(wb, "13_Decision_Register",
                      ["ID", "Decision", "Content"],
                      [[d["id"], d["title"], d["content"]] for d in ctx.get("decisions", [])],
                      doc, widths=(14, 36, 50))
    x.render_register(wb, "30_Traceability",
                      ["Source", "Target", "Relationship"],
                      [[t["source"], t["target"], t["relation"]] for t in ctx.get("trace_links", [])],
                      doc, widths=(24, 24, 20))
    x.render_register(wb, "31_Change_Request",
                      ["Code", "Title", "Requested Change", "Status"],
                      [[c["code"], c["title"], c["requested_change"], c["status"]] for c in ctx.get("change_requests", [])],
                      doc, widths=(12, 32, 46, 16))
    # TBD registers (no authoritative source yet)
    x.render_register(wb, "03_Milestones", ["Milestone", "Date", "Status", "Owner"], [], doc)
    x.render_register(wb, "04_Stakeholders", ["Name", "Organization", "Role", "Interest"], [], doc)
    x.render_register(wb, "14_Dependency_Register", ["ID", "Dependency", "Type", "Status"], [], doc)
    x.render_register(wb, "20_RAID", ["ID", "Type", "Description", "Owner", "Status"], [], doc)
    x.render_register(wb, "21_Action_Issue", ["ID", "Type", "Description", "Owner", "Status"], [], doc)
    x.render_register(wb, "40_Acceptance", ["ID", "Acceptance Item", "Status", "Date"], [], doc)
    x.render_register(wb, "41_Handover", ["Item", "Recipient", "Status", "Date"], [], doc)

    x.render_signoff(wb, doc)
    x.render_source_reference(wb, doc)
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_infra_workbook(ctx: dict, brand: str = "GEA_STANDARD") -> bytes:
    x = Xlsx(brand)
    wb = x.new_workbook()
    doc = _base_doc(ctx, "INFRA_DESIGN", f"{ctx['project']['name']} — Infrastructure Design")
    x.render_cover(wb, doc)
    x.render_document_control(wb, doc)
    x.render_revision_history(wb, doc)
    x.render_index(wb, doc, _index_entries(["00_Cover", "01_Document_Control", "02_Revision_History", "03_Sheet_Index",
                                            "10_Architecture_Summary", "11_Environment_Matrix", "12_Compute",
                                            "13_Storage", "14_Network", "15_IP_Plan", "16_Port_Firewall",
                                            "17_Load_Balancer", "18_Backup", "19_DR", "20_Monitoring",
                                            "21_Capacity", "22_Implementation", "23_Test", "24_AsBuilt",
                                            "90_Signoff", "99_Source"]))

    # Architecture summary from architecture diagrams
    arch_rows = []
    for d in ctx.get("architecture", []):
        for n in d["nodes"]:
            arch_rows.append([d["name"], n.semantic_id, n.name,
                              n.node_type.value if hasattr(n.node_type, "value") else str(n.node_type),
                              n.technology or "", n.environment or ""])
    x.render_register(wb, "10_Architecture_Summary",
                      ["Diagram", "Component ID", "Component", "Type", "Technology", "Environment"],
                      arch_rows, doc, widths=(24, 18, 34, 16, 20, 16))

    x.render_register(wb, "11_Environment_Matrix",
                      ["Environment", "Purpose", "Region / DC", "Network", "Compute", "Data", "HA", "Owner", "Status"],
                      [], doc)
    x.render_register(wb, "12_Compute", ["ID", "Host", "Role", "Spec", "Environment", "Status"], [], doc)
    x.render_register(wb, "13_Storage", ["ID", "Volume", "Type", "Capacity", "Environment", "Status"], [], doc)
    x.render_register(wb, "14_Network", ["ID", "Segment", "CIDR", "Purpose", "Environment", "Status"], [], doc)
    x.render_register(wb, "15_IP_Plan", ["IP / Range", "Purpose", "Environment", "Owner", "Status"], [], doc)
    x.render_register(wb, "16_Port_Firewall", ["Source", "Destination", "Protocol", "Port", "Direction", "Purpose", "Environment", "Status"], [], doc)
    x.render_register(wb, "17_Load_Balancer", ["ID", "Service", "Virtual IP", "Backends", "Environment", "Status"], [], doc)
    x.render_register(wb, "18_Backup", ["Asset", "Schedule", "Retention", "Tool", "Status"], [], doc)
    x.render_register(wb, "19_DR", ["Asset", "RPO", "RTO", "Mechanism", "Status"], [], doc)
    x.render_register(wb, "20_Monitoring", ["Asset", "Metric", "Threshold", "Alert", "Status"], [], doc)
    x.render_register(wb, "21_Capacity", ["Asset", "Current", "Projected", "Headroom", "Status"], [], doc)
    x.render_register(wb, "22_Implementation", ["Step", "Activity", "Owner", "Validation", "Status"], [], doc)
    x.render_register(wb, "23_Test", ["Test", "Expected Result", "Actual", "Status"], [], doc)
    x.render_register(wb, "24_AsBuilt", ["Component", "Configuration", "Environment", "Status"], [], doc)

    x.render_signoff(wb, doc)
    x.render_source_reference(wb, doc)
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_migration_workbook(ctx: dict, brand: str = "GEA_STANDARD") -> bytes:
    x = Xlsx(brand)
    wb = x.new_workbook()
    doc = _base_doc(ctx, "MIGRATION_PLAN", f"{ctx['project']['name']} — Migration Plan")
    x.render_cover(wb, doc)
    x.render_document_control(wb, doc)
    x.render_revision_history(wb, doc)
    x.render_index(wb, doc, _index_entries(["00_Cover", "01_Document_Control", "02_Revision_History", "03_Sheet_Index",
                                            "10_Assessment", "11_Source_Inventory", "12_Dependency_Matrix",
                                            "13_Readiness", "14_Wave_Plan", "15_Migration_Schedule", "16_Runbook",
                                            "17_Cutover", "18_Rollback", "19_Validation", "20_Reconciliation",
                                            "21_Hypercare", "22_Decommission", "90_Signoff", "99_Source"]))

    # flows become the runbook/wave content (honest: process flow steps)
    runbook_rows = []
    for f in ctx.get("flows", []):
        for s in f["steps"]:
            runbook_rows.append([f["name"], s.semantic_id, s.name,
                                 s.step_type.value if hasattr(s.step_type, "value") else str(s.step_type),
                                 s.position])
    x.render_register(wb, "16_Runbook",
                      ["Flow", "Step ID", "Step", "Type", "Position"],
                      runbook_rows, doc, widths=(26, 18, 36, 16, 10))

    x.render_register(wb, "10_Assessment", ["Item", "Finding", "Owner", "Status"], [], doc)
    x.render_register(wb, "11_Source_Inventory", ["System", "Type", "Location", "Owner", "Status"], [], doc)
    x.render_register(wb, "12_Dependency_Matrix", ["System", "Depends On", "Type", "Status"], [], doc)
    x.render_register(wb, "13_Readiness", ["System", "Check", "Result", "Status"], [], doc)
    x.render_register(wb, "14_Wave_Plan", ["Wave", "System", "Source", "Target", "Dependency", "Downtime", "Owner", "Planned Date", "Status"], [], doc)
    x.render_register(wb, "15_Migration_Schedule", ["Wave", "Activity", "Start", "End", "Status"], [], doc)
    x.render_register(wb, "17_Cutover", ["Step", "Phase", "Activity", "Owner", "Expected Result", "Validation", "Rollback Step", "Status"], [], doc)
    x.render_register(wb, "18_Rollback", ["Trigger", "Action", "Owner", "Status"], [], doc)
    x.render_register(wb, "19_Validation", ["Check", "Expected", "Actual", "Status"], [], doc)
    x.render_register(wb, "20_Reconciliation", ["Source", "Target", "Records", "Match", "Status"], [], doc)
    x.render_register(wb, "21_Hypercare", ["Item", "Window", "Owner", "Status"], [], doc)
    x.render_register(wb, "22_Decommission", ["System", "Action", "Date", "Status"], [], doc)

    x.render_signoff(wb, doc)
    x.render_source_reference(wb, doc)
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_security_workbook(ctx: dict, brand: str = "GEA_STANDARD") -> bytes:
    x = Xlsx(brand)
    wb = x.new_workbook()
    doc = _base_doc(ctx, "SECURITY_REGISTER", f"{ctx['project']['name']} — Security Register")
    x.render_cover(wb, doc)
    x.render_document_control(wb, doc)
    x.render_revision_history(wb, doc)
    x.render_index(wb, doc, _index_entries(["00_Cover", "01_Document_Control", "02_Revision_History", "03_Sheet_Index",
                                            "10_Security_Requirements", "11_Risk_Assessment", "12_Threat_Model",
                                            "13_IAM", "14_Access_Matrix", "15_Network_Security", "16_Hardening",
                                            "17_Compliance", "18_Vulnerability", "19_Remediation", "20_Security_Test",
                                            "21_Acceptance", "90_Signoff", "99_Source"]))
    x.render_register(wb, "10_Security_Requirements", ["ID", "Requirement", "Source", "Status"], [], doc)
    x.render_register(wb, "11_Risk_Assessment", ["Risk ID", "Asset", "Threat", "Vulnerability", "Likelihood", "Impact", "Rating", "Mitigation", "Owner", "Status"], [], doc)
    x.render_register(wb, "12_Threat_Model", ["Asset", "Threat", "Vector", "Impact", "Status"], [], doc)
    x.render_register(wb, "13_IAM", ["Role", "Privilege", "Environment", "Status"], [], doc)
    x.render_register(wb, "14_Access_Matrix", ["Role", "Resource", "Access", "Status"], [], doc)
    x.render_register(wb, "15_Network_Security", ["Zone", "Control", "Purpose", "Status"], [], doc)
    x.render_register(wb, "16_Hardening", ["Asset", "Control", "Applied", "Status"], [], doc)
    x.render_register(wb, "17_Compliance", ["Requirement", "Control", "Evidence", "Status"], [], doc)
    x.render_register(wb, "18_Vulnerability", ["ID", "Asset", "Finding", "Severity", "Status"], [], doc)
    x.render_register(wb, "19_Remediation", ["ID", "Action", "Owner", "Due", "Status"], [], doc)
    x.render_register(wb, "20_Security_Test", ["Test", "Expected", "Actual", "Status"], [], doc)
    x.render_register(wb, "21_Acceptance", ["Item", "Result", "Date", "Status"], [], doc)
    x.render_signoff(wb, doc)
    x.render_source_reference(wb, doc)
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_application_workbook(ctx: dict, brand: str = "GEA_STANDARD") -> bytes:
    x = Xlsx(brand)
    wb = x.new_workbook()
    doc = _base_doc(ctx, "APPLICATION_DESIGN", f"{ctx['project']['name']} — Application Design")
    x.render_cover(wb, doc)
    x.render_document_control(wb, doc)
    x.render_revision_history(wb, doc)
    x.render_index(wb, doc, _index_entries(["00_Cover", "01_Document_Control", "02_Revision_History", "03_Sheet_Index",
                                            "10_Requirements", "11_Function_List", "12_NFR", "13_Application_Architecture",
                                            "14_Screen_List", "15_API", "16_Integration", "17_Data_Model", "18_Batch_Jobs",
                                            "19_Error_Handling", "20_Deployment", "21_Test", "22_UAT",
                                            "90_Signoff", "99_Source"]))
    x.render_register(wb, "10_Requirements",
                      ["Code", "Title", "Status", "Priority", "Source"],
                      [[r["code"], r["title"], r["status"], r["priority"], r["source"]] for r in ctx.get("requirements", [])],
                      doc, widths=(14, 52, 12, 10, 18))
    x.render_register(wb, "11_Function_List", ["Function ID", "Module", "Function", "Description", "Actor", "Priority", "Requirement Ref", "Status"], [], doc)
    x.render_register(wb, "12_NFR", ["ID", "Requirement", "Category", "Target", "Status"], [], doc)
    x.render_register(wb, "13_Application_Architecture", ["Component", "Type", "Technology", "Purpose"], [], doc)
    x.render_register(wb, "14_Screen_List", ["Screen", "Module", "Purpose", "Status"], [], doc)
    x.render_register(wb, "15_API", ["Method", "Path", "Purpose", "Status"], [], doc)
    x.render_register(wb, "16_Integration", ["Interface", "Direction", "Protocol", "Status"], [], doc)
    x.render_register(wb, "17_Data_Model", ["Entity", "Fields", "Relationship", "Status"], [], doc)
    x.render_register(wb, "18_Batch_Jobs", ["Job", "Schedule", "Purpose", "Status"], [], doc)
    x.render_register(wb, "19_Error_Handling", ["Code", "Scenario", "Handling", "Status"], [], doc)
    x.render_register(wb, "20_Deployment", ["Environment", "Component", "Version", "Status"], [], doc)
    x.render_register(wb, "21_Test", ["Test", "Expected", "Actual", "Status"], [], doc)
    x.render_register(wb, "22_UAT", ["Scenario", "Result", "Date", "Status"], [], doc)
    x.render_signoff(wb, doc)
    x.render_source_reference(wb, doc)
    wb.active = 0
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


WORKBOOK_BUILDERS = {
    "PROJECT_MASTER": build_master_workbook,
    "INFRA_DESIGN": build_infra_workbook,
    "MIGRATION_PLAN": build_migration_workbook,
    "SECURITY_REGISTER": build_security_workbook,
    "APPLICATION_DESIGN": build_application_workbook,
}


def build_workbook(db: Session, project: m.Project, mode: str, brand: str = "GEA_STANDARD") -> bytes:
    if mode not in WORKBOOK_BUILDERS:
        raise ValueError(f"Unknown workbook mode: {mode}")
    ctx = build_context(db, project)
    return WORKBOOK_BUILDERS[mode](ctx, brand)
